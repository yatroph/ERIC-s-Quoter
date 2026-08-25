from __future__ import annotations

from dataclasses import replace

from openpyxl import load_workbook
import pytest

from erics_quoter.labour_rates import (
    DEFAULT_LABOUR_RATE_SETTINGS,
    LabourRateSettingsError,
    LabourRateSettings,
    labour_rates_path,
    load_labour_rate_settings,
    save_labour_rate_settings,
)
from erics_quoter.models import QuoteMode, QuoteRequest
from erics_quoter.workbook_service import WorkbookGenerationError, generate_quote


def _custom_settings() -> LabourRateSettings:
    rates = list(DEFAULT_LABOUR_RATE_SETTINGS.rates)
    rates[0] = replace(rates[0], regular=111.11)
    return replace(
        DEFAULT_LABOUR_RATE_SETTINGS,
        year=2027,
        truck_premium=19.25,
        sell_margin=0.78,
        rates=tuple(rates),
    )


def test_labour_rates_are_saved_in_local_app_data(tmp_path, monkeypatch):
    monkeypatch.setenv("LOCALAPPDATA", str(tmp_path))
    custom = _custom_settings()

    assert load_labour_rate_settings() == DEFAULT_LABOUR_RATE_SETTINGS
    save_labour_rate_settings(custom)

    assert labour_rates_path().is_file()
    assert load_labour_rate_settings() == custom


def test_invalid_saved_labour_rates_are_never_silently_used(tmp_path, monkeypatch):
    monkeypatch.setenv("LOCALAPPDATA", str(tmp_path))
    path = labour_rates_path()
    path.parent.mkdir(parents=True)
    path.write_text('{"year": "not a year"}', encoding="utf-8")

    with pytest.raises(LabourRateSettingsError, match="Review and save"):
        load_labour_rate_settings()


def test_unreadable_encoding_is_reported_instead_of_using_defaults(
    tmp_path,
    monkeypatch,
):
    monkeypatch.setenv("LOCALAPPDATA", str(tmp_path))
    path = labour_rates_path()
    path.parent.mkdir(parents=True)
    path.write_bytes(b"\xff\xfe\xfa")

    with pytest.raises(LabourRateSettingsError, match="could not be read"):
        load_labour_rate_settings()


def test_custom_labour_rates_flow_into_workbook_and_leave_bonding_intact(tmp_path):
    destination = generate_quote(
        QuoteRequest(
            mode=QuoteMode.MULTI_YEAR,
            customer_name="Custom Rates",
            contract_years=1,
            output_directory=tmp_path,
            labour_rates=_custom_settings(),
        )
    )

    workbook = load_workbook(destination, data_only=False)
    rates = workbook["Labour Rates"]
    pricing = workbook["Pricing"]
    assert rates["A1"].value.endswith("2027")
    assert rates["B6"].value == 111.11
    assert rates["B19"].value == 19.25
    assert rates["B20"].value == 0.78
    assert pricing["C73"].value == "='Labour Rates'!$B$6"
    assert pricing["B92"].value == "Bonding"
    assert pricing["C93"].value == 8.7


def test_generation_uses_persisted_rates_when_request_does_not_override_them(
    tmp_path,
    monkeypatch,
):
    monkeypatch.setenv("LOCALAPPDATA", str(tmp_path / "app-data"))
    save_labour_rate_settings(_custom_settings())

    destination = generate_quote(
        QuoteRequest(
            mode=QuoteMode.MULTI_YEAR,
            customer_name="Persisted Rates",
            contract_years=1,
            output_directory=tmp_path / "output",
        )
    )

    workbook = load_workbook(destination, data_only=False)
    rates = workbook["Labour Rates"]
    assert rates["A1"].value.endswith("2027")
    assert rates["B6"].value == 111.11


def test_generation_rejects_corrupted_persisted_rates(tmp_path, monkeypatch):
    monkeypatch.setenv("LOCALAPPDATA", str(tmp_path / "app-data"))
    path = labour_rates_path()
    path.parent.mkdir(parents=True)
    path.write_text("not json", encoding="utf-8")

    with pytest.raises(WorkbookGenerationError, match="Saved labour rates are invalid"):
        generate_quote(
            QuoteRequest(
                mode=QuoteMode.MULTI_YEAR,
                customer_name="Unsafe Rates",
                contract_years=1,
                output_directory=tmp_path / "output",
            )
        )


def test_invalid_custom_labour_rates_are_rejected(tmp_path):
    invalid = replace(DEFAULT_LABOUR_RATE_SETTINGS, sell_margin=0)

    with pytest.raises(WorkbookGenerationError, match="sell margin"):
        generate_quote(
            QuoteRequest(
                mode=QuoteMode.MULTI_YEAR,
                customer_name="Invalid Rates",
                contract_years=1,
                output_directory=tmp_path,
                labour_rates=invalid,
            )
        )
