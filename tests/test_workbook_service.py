from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor
from threading import Barrier
import warnings
from zipfile import ZipFile

from openpyxl import load_workbook
import pytest

from erics_quoter import workbook_service
from erics_quoter.history import load_recent_customers, remember_customer
from erics_quoter.models import LocationSpec, QuoteMode, QuoteRequest
from erics_quoter.workbook_service import (
    WorkbookGenerationError,
    _shift_formula_rows,
    generate_quote,
)


def test_structural_formula_shift_only_moves_local_references():
    formula = "=SUM(A31,A32,'Other sheet'!P36,$K$70,LOG10(A1))"

    assert _shift_formula_rows(formula, insertion_row=32, amount=4) == (
        "=SUM(A31,A36,'Other sheet'!P36,$K$74,LOG10(A1))"
    )


def test_generates_multi_year_copy_with_customer_and_full_print_area(tmp_path):
    destination = generate_quote(
        QuoteRequest(
            mode=QuoteMode.MULTI_YEAR,
            customer_name="Northwind Foods",
            project_scope="Service agreement",
            output_directory=tmp_path,
        )
    )

    assert destination.exists()
    workbook = load_workbook(destination, data_only=False)
    sheet = workbook["Multi Year Contract Pricing"]
    assert sheet["A1"].value == (
        "Customer and Project Scope - Northwind Foods | Service agreement"
    )
    assert sheet["Q8"].value == "=P8*$Q$6"
    assert sheet.print_area == "'Multi Year Contract Pricing'!$A$1:$V$80"
    assert workbook["Pricing"].sheet_state == "hidden"
    assert workbook["Pricing"]["R37"].value == "=IFERROR(M37/P37,0)"
    assert workbook["Pricing"]["R51"].value == "=IFERROR(M51/P51,0)"
    assert workbook.active.title == "Multi Year Contract Pricing"
    assert sheet.sheet_view.tabSelected is True


def test_one_year_uses_standard_pricing_sheet(tmp_path):
    destination = generate_quote(
        QuoteRequest(
            mode=QuoteMode.MULTI_YEAR,
            customer_name="One Year Customer",
            project_scope="Standard pricing",
            contract_years=1,
            output_directory=tmp_path,
        )
    )

    workbook = load_workbook(destination, data_only=False)
    pricing = workbook["Pricing"]
    assert destination.name.endswith(" - Pricing.xlsx")
    assert pricing["A1"].value == (
        "Customer and Project Scope - One Year Customer | Standard pricing"
    )
    assert pricing.print_area == "'Pricing'!$A$1:$P$65"
    assert pricing["R27"].value == "=IFERROR((N27+H27)/P27,0)"
    assert pricing.sheet_state == "visible"
    assert workbook["Multi Year Contract Pricing"].sheet_state == "hidden"
    assert workbook.active.title == "Pricing"
    assert pricing.sheet_view.tabSelected is True


@pytest.mark.parametrize("contract_years", [2, 3, 4, 5])
def test_multi_year_hides_unused_years_and_limits_totals(tmp_path, contract_years):
    destination = generate_quote(
        QuoteRequest(
            mode=QuoteMode.MULTI_YEAR,
            customer_name="Variable Term Customer",
            contract_years=contract_years,
            output_directory=tmp_path,
        )
    )

    workbook = load_workbook(destination, data_only=False)
    sheet = workbook["Multi Year Contract Pricing"]
    year_columns = ("P", "Q", "R", "S", "T")
    for year_number, column in enumerate(year_columns, start=1):
        assert sheet.column_dimensions[column].hidden is (
            year_number > contract_years
        )
    last_year_column = year_columns[contract_years - 1]
    assert sheet["V69"].value == f"=SUM(P69:{last_year_column}69)"
    assert sheet["V79"].value == f"=SUM(P79:{last_year_column}79)"
    assert workbook["Pricing"].sheet_state == "hidden"
    assert workbook.active.title == "Multi Year Contract Pricing"
    assert destination.name.endswith(f" - {contract_years} Year Contract.xlsx")


@pytest.mark.parametrize("contract_years", [0, 6])
def test_rejects_contract_term_outside_one_to_five_years(
    tmp_path,
    contract_years,
):
    with pytest.raises(
        WorkbookGenerationError,
        match="between 1 and 5 contract years",
    ):
        generate_quote(
            QuoteRequest(
                mode=QuoteMode.MULTI_YEAR,
                customer_name="Invalid Term",
                contract_years=contract_years,
                output_directory=tmp_path,
            )
        )


def test_generates_required_and_optional_location_sheets(tmp_path):
    destination = generate_quote(
        QuoteRequest(
            mode=QuoteMode.MULTI_LOCATION,
            customer_name="Contoso",
            project_scope="Plant retrofit",
            locations=(
                LocationSpec("Main plant"),
                LocationSpec("Warehouse option", optional=True),
                LocationSpec("Admin building"),
                LocationSpec("Future line", optional=True),
            ),
            output_directory=tmp_path,
        )
    )

    workbook = load_workbook(destination, data_only=False)
    summary = workbook["Cost Summary"]
    assert summary["A1"].value == "Customer and Project Scope - Contoso | Plant retrofit"
    assert workbook["Main plant"]["R1"].value == "Main plant"
    assert workbook["Admin building"]["R1"].value == "Admin building"
    assert workbook["Warehouse option (Optional)"]["R1"].value == (
        "Warehouse option (Optional)"
    )
    assert workbook["Future line (Optional)"]["R1"].value == (
        "Future line (Optional)"
    )
    assert summary["B21"].value == "='Main plant'!R1"
    assert summary["B22"].value == "='Admin building'!R1"
    assert summary["B30"].value == "='Warehouse option (Optional)'!R1"
    assert summary["B31"].value == "='Future line (Optional)'!R1"
    assert summary["J32"].value == "=SUM(J30:J31)"
    assert summary["H30"].value == "=F30*G30"
    assert summary["J35"].value == "=J18+J27"
    assert summary["J39"].value == 0
    assert summary["J43"].value == "=(J32)"
    assert summary["J47"].value == "=J45+J46"
    assert summary["J52"].value == "Price"
    assert summary["A27"].value == "Subtotal Required Locations"
    assert summary["B29"].value == "Optional Location"
    assert summary["A32"].value == "Subtotal Optional Locations"
    assert summary.print_area == "'Cost Summary'!$A$1:$J$32"
    assert "Location 3" not in workbook.sheetnames
    assert summary.sheet_view.tabSelected is True
    assert workbook["Main plant"].sheet_view.tabSelected is False
    assert summary["D61"].value == "='Labour Rates'!$E$6"
    assert summary["C70"].value == "='Labour Rates'!$I$6"
    assert summary["I86"].value == "=I85/0.9"
    assert summary["K70"].value == "='Labour Rates'!$H$6"
    assert str(summary.data_validations.dataValidation[0].sqref) == (
        "I6:I12 I16 K70"
    )
    assert summary.row_dimensions[60].height == 32.25
    assert summary.row_dimensions[69].height == 32.25
    assert workbook["Future line (Optional)"].print_area == (
        "'Future line (Optional)'!$A$1:$P$62"
    )


def test_supports_six_optional_locations_without_overwriting(tmp_path):
    request = QuoteRequest(
        mode=QuoteMode.MULTI_LOCATION,
        customer_name="Fabrikam",
        locations=tuple(
            LocationSpec(f"Optional site {index}", optional=True)
            for index in range(1, 7)
        ),
        output_directory=tmp_path,
    )

    first = generate_quote(request)
    second = generate_quote(request)
    assert first != second

    workbook = load_workbook(first, data_only=False)
    summary = workbook["Cost Summary"]
    assert workbook["Optional site 6 (Optional)"]["R1"].value == (
        "Optional site 6 (Optional)"
    )
    assert summary["B35"].value == "='Optional site 6 (Optional)'!R1"
    assert summary.row_dimensions[35].hidden is False
    assert not any(name.startswith("Location ") for name in workbook.sheetnames)


def test_formula_looking_location_name_is_stored_as_literal_text(tmp_path):
    payload = '=HYPERLINK("https://example.invalid","Open")'
    destination = generate_quote(
        QuoteRequest(
            mode=QuoteMode.MULTI_LOCATION,
            customer_name="Formula Test",
            locations=(LocationSpec(payload),),
            output_directory=tmp_path,
        )
    )

    workbook = load_workbook(destination, data_only=False)
    sheet = next(
        worksheet
        for worksheet in workbook.worksheets
        if worksheet.title not in {"Cost Summary", "Labour Rates"}
    )
    assert sheet["A1"].value == payload
    assert sheet["A1"].data_type == "s"
    assert sheet["R1"].value == payload
    assert sheet["R1"].data_type == "s"


def test_supports_twenty_locations_and_scales_summary(tmp_path):
    locations = tuple(
        LocationSpec(f"Site {index}", optional=index > 12)
        for index in range(1, 21)
    )
    destination = generate_quote(
        QuoteRequest(
            mode=QuoteMode.MULTI_LOCATION,
            customer_name="Twenty Site Customer",
            locations=locations,
            output_directory=tmp_path,
        )
    )

    workbook = load_workbook(destination, data_only=False)
    summary = workbook["Cost Summary"]
    assert len(workbook.sheetnames) == 22
    assert workbook.sheetnames == (
        ["Cost Summary"]
        + [f"Site {index}" for index in range(1, 13)]
        + [f"Site {index} (Optional)" for index in range(13, 21)]
        + ["Labour Rates"]
    )
    assert summary["B32"].value == "='Site 12'!R1"
    assert summary["A33"].value == "Subtotal Required Locations"
    assert summary["B43"].value == "='Site 20 (Optional)'!R1"
    assert summary["A44"].value == "Subtotal Optional Locations"
    assert summary["J44"].value == "=SUM(J36:J43)"
    assert summary["J55"].value == "=(J44)"
    assert summary["J59"].value == "=J57+J58"
    assert summary.print_area == "'Cost Summary'!$A$1:$J$44"
    assert summary["K82"].value == "='Labour Rates'!$H$6"


@pytest.mark.parametrize("contract_years", [1, 3, 5])
def test_contract_pricing_can_create_named_location_tabs(tmp_path, contract_years):
    destination = generate_quote(
        QuoteRequest(
            mode=QuoteMode.MULTI_YEAR,
            customer_name="Contract Sites",
            project_scope="Regional service",
            locations=(
                LocationSpec("Toronto/DC"),
                LocationSpec("Future Campus", optional=True),
            ),
            contract_years=contract_years,
            include_locations=True,
            output_directory=tmp_path,
        )
    )

    workbook = load_workbook(destination, data_only=False)
    assert workbook.sheetnames == [
        "Toronto-DC",
        "Future Campus (Optional)",
        "Labour Rates",
    ]
    expected_display_names = ("Toronto/DC", "Future Campus (Optional)")
    for title, display_name in zip(
        workbook.sheetnames[:2],
        expected_display_names,
        strict=True,
    ):
        sheet = workbook[title]
        assert "Contract Sites | Regional service" in sheet["A1"].value
        assert sheet["A1"].value.endswith(f"Location - {display_name}")
        assert sheet.sheet_state == "visible"
        if contract_years == 1:
            assert sheet.print_area == f"'{title}'!$A$1:$P$65"
        else:
            assert sheet.print_area == f"'{title}'!$A$1:$V$80"
            assert sheet["V79"].value == (
                f"=SUM(P79:{('P', 'Q', 'R', 'S', 'T')[contract_years - 1]}79)"
            )
    assert workbook.active.title == "Toronto-DC"


def test_labour_rates_are_global_and_bonding_is_preserved(tmp_path):
    destination = generate_quote(
        QuoteRequest(
            mode=QuoteMode.MULTI_YEAR,
            customer_name="Rates Customer",
            contract_years=1,
            output_directory=tmp_path,
        )
    )

    workbook = load_workbook(destination, data_only=False)
    rates = workbook["Labour Rates"]
    pricing = workbook["Pricing"]
    assert rates["A1"].value == "GDI Ainsworth Labour Rates — 2026"
    assert rates["A6"].value == "FOREMAN"
    assert rates["B6"].value == 102.0
    assert rates["B19"].value == 18.5
    assert rates["B20"].value == 0.8
    assert pricing["C73"].value == "='Labour Rates'!$B$6"
    assert pricing["D73"].value == "='Labour Rates'!$E$6"
    assert pricing["C83"].value == "='Labour Rates'!$I$6"
    assert pricing["K83"].value == "='Labour Rates'!$H$6"
    assert pricing["B92"].value == "Bonding"
    assert pricing["C93"].value == 8.7
    assert pricing["I99"].value == "=I98/0.9"


def test_location_sheet_titles_are_unique_and_excel_safe(tmp_path):
    destination = generate_quote(
        QuoteRequest(
            mode=QuoteMode.MULTI_LOCATION,
            customer_name="Sheet Names",
            locations=(
                LocationSpec("North/Plant"),
                LocationSpec("North:Plant"),
                LocationSpec("Labour Rates"),
                LocationSpec("A" * 80, optional=True),
            ),
            output_directory=tmp_path,
        )
    )

    workbook = load_workbook(destination, data_only=False)
    assert workbook.sheetnames == [
        "Cost Summary",
        "North-Plant",
        "North-Plant (2)",
        "Labour Rates (2)",
        "AAAAAAAAAAAAAAAAAAAA (Optional)",
        "Labour Rates",
    ]
    assert all(len(title) <= 31 for title in workbook.sheetnames)


def test_rejects_more_than_twenty_locations(tmp_path):
    with pytest.raises(WorkbookGenerationError, match="between 1 and 20 locations"):
        generate_quote(
            QuoteRequest(
                mode=QuoteMode.MULTI_LOCATION,
                customer_name="Too Many Sites",
                locations=tuple(LocationSpec(str(index)) for index in range(21)),
                output_directory=tmp_path,
            )
        )


def test_twenty_location_contract_generation_emits_no_excel_title_warnings(tmp_path):
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        destination = generate_quote(
            QuoteRequest(
                mode=QuoteMode.MULTI_YEAR,
                customer_name="Twenty Contract Sites",
                locations=tuple(
                    LocationSpec(f"Contract Site {index}")
                    for index in range(1, 21)
                ),
                contract_years=5,
                include_locations=True,
                output_directory=tmp_path,
            )
        )

    assert caught == []
    workbook = load_workbook(destination, data_only=False)
    assert len(workbook.sheetnames) == 21
    assert workbook.sheetnames[0] == "Contract Site 1"
    assert workbook.sheetnames[-2] == "Contract Site 20"
    assert workbook.sheetnames[-1] == "Labour Rates"


def test_apostrophe_in_location_name_is_escaped_in_summary_formula(tmp_path):
    destination = generate_quote(
        QuoteRequest(
            mode=QuoteMode.MULTI_LOCATION,
            customer_name="Quoted Sheet Name",
            locations=(LocationSpec("King's Site"),),
            output_directory=tmp_path,
        )
    )

    workbook = load_workbook(destination, data_only=False)
    assert "King's Site" in workbook.sheetnames
    assert workbook["Cost Summary"]["B21"].value == "='King''s Site'!R1"


def test_location_sheet_titles_respect_excels_utf16_unit_limit(tmp_path):
    destination = generate_quote(
        QuoteRequest(
            mode=QuoteMode.MULTI_LOCATION,
            customer_name="Unicode Sheet Name",
            locations=(
                LocationSpec("😀" * 31),
                LocationSpec("🚀" * 31, optional=True),
            ),
            output_directory=tmp_path,
        )
    )

    workbook = load_workbook(destination, data_only=False)
    site_titles = [
        title
        for title in workbook.sheetnames
        if title not in {"Cost Summary", "Labour Rates"}
    ]
    assert site_titles == ["😀" * 15, f"{'🚀' * 10} (Optional)"]
    assert all(len(title.encode("utf-16-le")) // 2 <= 31 for title in site_titles)


@pytest.mark.parametrize(
    ("customer_name", "location_name", "message"),
    [
        ("Bad\x00Customer", "Location 1", "unsupported character"),
        ("Customer", "Bad\x00Location", "unsupported character"),
        ("C" * 501, "Location 1", "500 characters or fewer"),
        ("Customer", "L" * 501, "500 characters or fewer"),
    ],
)
def test_rejects_input_excel_cannot_safely_store(
    tmp_path,
    customer_name,
    location_name,
    message,
):
    with pytest.raises(WorkbookGenerationError, match=message):
        generate_quote(
            QuoteRequest(
                mode=QuoteMode.MULTI_LOCATION,
                customer_name=customer_name,
                locations=(LocationSpec(location_name),),
                output_directory=tmp_path,
            )
        )


@pytest.mark.parametrize("contract_years", [True, 1.5, "2"])
def test_rejects_non_integer_contract_terms(tmp_path, contract_years):
    with pytest.raises(WorkbookGenerationError, match="whole number"):
        generate_quote(
            QuoteRequest(
                mode=QuoteMode.MULTI_YEAR,
                customer_name="Invalid Term Type",
                contract_years=contract_years,
                output_directory=tmp_path,
            )
        )


def test_existing_file_cannot_be_used_as_output_folder(tmp_path):
    output_file = tmp_path / "not-a-folder"
    output_file.write_text("sentinel", encoding="utf-8")

    with pytest.raises(WorkbookGenerationError, match="output folder"):
        generate_quote(
            QuoteRequest(
                mode=QuoteMode.MULTI_YEAR,
                customer_name="Output Error",
                contract_years=1,
                output_directory=output_file,
            )
        )
    assert output_file.read_text(encoding="utf-8") == "sentinel"


def test_corrupted_template_uses_friendly_error_path(tmp_path, monkeypatch):
    fake_root = tmp_path / "resources"
    costing = fake_root / "Costing sheets"
    costing.mkdir(parents=True)
    (costing / workbook_service.MULTI_YEAR_TEMPLATE.name).write_bytes(b"not a zip")
    monkeypatch.setattr(workbook_service, "resource_root", lambda: fake_root)

    with pytest.raises(WorkbookGenerationError, match="Could not create"):
        generate_quote(
            QuoteRequest(
                mode=QuoteMode.MULTI_YEAR,
                customer_name="Corrupt Template",
                contract_years=1,
                output_directory=tmp_path / "output",
            )
        )


def test_concurrent_generations_reserve_unique_destinations(tmp_path):
    request = QuoteRequest(
        mode=QuoteMode.MULTI_YEAR,
        customer_name="Concurrent Customer",
        contract_years=1,
        output_directory=tmp_path,
    )
    workers = 6
    barrier = Barrier(workers)

    def create_after_barrier():
        barrier.wait(timeout=30)
        return generate_quote(request)

    with ThreadPoolExecutor(max_workers=workers) as executor:
        destinations = list(executor.map(lambda _index: create_after_barrier(), range(workers)))

    assert len(set(destinations)) == workers
    assert all(destination.exists() for destination in destinations)


def test_customer_history_survives_concurrent_writers(tmp_path, monkeypatch):
    monkeypatch.setenv("LOCALAPPDATA", str(tmp_path))
    customers = [f"Customer {index}" for index in range(20)]

    with ThreadPoolExecutor(max_workers=20) as executor:
        list(executor.map(remember_customer, customers))

    saved = load_recent_customers()
    assert len(saved) == 12
    assert len(set(saved)) == 12
    assert set(saved).issubset(customers)


def test_source_templates_do_not_publish_printer_blobs_or_personal_authors():
    for template_name in (
        workbook_service.MULTI_LOCATION_TEMPLATE,
        workbook_service.MULTI_YEAR_TEMPLATE,
    ):
        template = workbook_service.resource_root() / template_name
        with ZipFile(template) as archive:
            assert not any("printerSettings" in name for name in archive.namelist())
        workbook = load_workbook(template, read_only=True)
        assert workbook.properties.creator == "GDI Ainsworth"
        assert workbook.properties.lastModifiedBy == "GDI Ainsworth"
        workbook.close()
