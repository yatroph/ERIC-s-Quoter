"""Generate quote workbooks while preserving the supplied Excel templates."""

from __future__ import annotations

import os
import re
import sys
from copy import copy, deepcopy
from pathlib import Path
from tempfile import NamedTemporaryFile

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .labour_rates import (
    LabourRateSettings,
    load_labour_rate_settings,
    validate_labour_rate_settings,
)
from .models import LocationSpec, QuoteMode, QuoteRequest

MULTI_LOCATION_TEMPLATE = Path("Costing sheets") / "Costing Sheet Multiple Location.2026.xlsx"
MULTI_YEAR_TEMPLATE = Path("Costing sheets") / "Costing Sheet.template 2026.xlsx"
MAX_LOCATIONS = 20
MAX_CONTRACT_YEARS = 5
MAX_INPUT_CHARACTERS = 500
BASE_REQUIRED_ROW_COUNT = 6
BASE_OPTION_ROW_COUNT = 2
REQUIRED_START_ROW = 21
BASE_REQUIRED_SUBTOTAL_ROW = 27
BASE_OPTION_START_ROW = 30
BASE_OPTION_SUBTOTAL_ROW = 32
LABOUR_RATES_SHEET = "Labour Rates"
INVALID_SHEET_TITLE = re.compile(r"[\\/*?:\[\]]")
ILLEGAL_EXCEL_TEXT = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")
CELL_REFERENCE_PATTERN = re.compile(
    r"(?<![A-Za-z0-9_!.])(?P<column>\$?[A-Z]{1,3})(?P<absolute_row>\$?)(?P<row>\d+)"
    r"(?![A-Za-z0-9_]|\s*\()"
)


class WorkbookGenerationError(RuntimeError):
    """Raised when a user request cannot produce a safe workbook."""


def resource_root() -> Path:
    frozen_root = getattr(sys, "_MEIPASS", None)
    if frozen_root:
        return Path(frozen_root)
    return Path(__file__).resolve().parents[1]


def generate_quote(request: QuoteRequest) -> Path:
    normalized = _validate_request(request)
    output_directory = normalized.output_directory.expanduser().resolve()
    try:
        output_directory.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        raise WorkbookGenerationError(f"Could not use the output folder: {exc}") from exc

    template_name = (
        MULTI_LOCATION_TEMPLATE
        if normalized.mode is QuoteMode.MULTI_LOCATION
        else MULTI_YEAR_TEMPLATE
    )
    template_path = resource_root() / template_name
    if not template_path.is_file():
        raise WorkbookGenerationError(f"Template not found: {template_path}")

    workbook = None
    temp_path: Path | None = None
    destination: Path | None = None
    committed = False
    try:
        workbook = load_workbook(template_path, data_only=False, keep_links=True)
        if normalized.mode is QuoteMode.MULTI_LOCATION:
            _prepare_multi_location(workbook, normalized)
        elif normalized.include_locations:
            _prepare_multi_year_locations(workbook, normalized)
        else:
            _prepare_multi_year(workbook, normalized)

        if normalized.labour_rates is None:
            raise WorkbookGenerationError("Labour-rate settings were not available.")
        _add_labour_rates_reference(workbook, normalized.labour_rates)
        _link_labour_rate_tables(workbook, normalized.labour_rates)
        _sanitize_workbook_metadata(workbook)
        _request_full_recalculation(workbook)
        with NamedTemporaryFile(
            prefix="erics-quoter-",
            suffix=".xlsx",
            dir=output_directory,
            delete=False,
        ) as temp_file:
            temp_path = Path(temp_file.name)
        workbook.save(temp_path)
        destination = _unique_destination(normalized, output_directory)
        temp_path.replace(destination)
        committed = True
    except Exception as exc:
        if temp_path is not None:
            temp_path.unlink(missing_ok=True)
        if destination is not None and not committed:
            try:
                if destination.stat().st_size == 0:
                    destination.unlink(missing_ok=True)
            except OSError:
                pass
        if isinstance(exc, WorkbookGenerationError):
            raise
        raise WorkbookGenerationError(f"Could not create the workbook: {exc}") from exc
    finally:
        if workbook is not None:
            workbook.close()

    if destination is None:
        raise WorkbookGenerationError("Could not reserve a destination workbook name.")
    return destination


def _validate_request(request: QuoteRequest) -> QuoteRequest:
    if not isinstance(request.mode, QuoteMode):
        raise WorkbookGenerationError("Choose a valid quote mode.")

    customer = _validated_text(request.customer_name, "customer name", required=True)
    project = _validated_text(request.project_scope, "project description")

    try:
        raw_locations = tuple(request.locations)
    except TypeError as exc:
        raise WorkbookGenerationError("Locations must be a list of location entries.") from exc

    locations: list[LocationSpec] = []
    for location in raw_locations:
        if not isinstance(location, LocationSpec):
            raise WorkbookGenerationError("Every location must be a valid location entry.")
        locations.append(
            LocationSpec(
                _validated_text(location.name, "location name", required=True),
                bool(location.optional),
            )
        )
    normalized_locations = tuple(locations)
    include_locations = bool(request.include_locations)
    needs_locations = request.mode is QuoteMode.MULTI_LOCATION or include_locations
    if needs_locations:
        if not 1 <= len(normalized_locations) <= MAX_LOCATIONS:
            raise WorkbookGenerationError(
                f"Choose between 1 and {MAX_LOCATIONS} locations."
            )
    if request.mode is QuoteMode.MULTI_YEAR:
        if isinstance(request.contract_years, bool) or not isinstance(
            request.contract_years, int
        ):
            raise WorkbookGenerationError("Contract years must be a whole number.")
        if not 1 <= request.contract_years <= MAX_CONTRACT_YEARS:
            raise WorkbookGenerationError("Choose between 1 and 5 contract years.")

    try:
        output_directory = Path(request.output_directory)
    except TypeError as exc:
        raise WorkbookGenerationError("Choose a valid output folder.") from exc

    try:
        labour_rates = validate_labour_rate_settings(
            request.labour_rates
            if request.labour_rates is not None
            else load_labour_rate_settings()
        )
    except ValueError as exc:
        raise WorkbookGenerationError(str(exc)) from exc

    return QuoteRequest(
        mode=request.mode,
        customer_name=customer,
        project_scope=project,
        locations=normalized_locations,
        contract_years=request.contract_years,
        include_locations=include_locations,
        output_directory=output_directory,
        labour_rates=labour_rates,
    )


def _validated_text(value: object, label: str, *, required: bool = False) -> str:
    if not isinstance(value, str):
        raise WorkbookGenerationError(f"Enter a valid {label}.")
    normalized = value.strip()
    if required and not normalized:
        if label == "customer name":
            raise WorkbookGenerationError("Enter a customer name.")
        raise WorkbookGenerationError("Every location needs a name.")
    if len(normalized) > MAX_INPUT_CHARACTERS:
        raise WorkbookGenerationError(
            f"Keep the {label} to {MAX_INPUT_CHARACTERS} characters or fewer."
        )
    if ILLEGAL_EXCEL_TEXT.search(normalized):
        raise WorkbookGenerationError(f"The {label} contains an unsupported character.")
    return normalized


def _prepare_multi_year(workbook, request: QuoteRequest):
    title = _customer_title(request)
    for sheet_name in ("Pricing", "Multi Year Contract Pricing"):
        worksheet = workbook[sheet_name]
        _set_title_cell(worksheet, title)

    pricing = workbook["Pricing"]
    multi_year = workbook["Multi Year Contract Pricing"]
    pricing["R27"] = "=IFERROR((N27+H27)/P27,0)"
    pricing["R37"] = "=IFERROR(M37/P37,0)"
    pricing["R51"] = "=IFERROR(M51/P51,0)"
    if request.contract_years == 1:
        pricing.sheet_state = "visible"
        multi_year.sheet_state = "hidden"
        pricing.print_area = "A1:P65"
        pricing.sheet_properties.pageSetUpPr.fitToPage = True
        pricing.page_setup.fitToWidth = 1
        pricing.page_setup.fitToHeight = 1
        _select_only(workbook, pricing)
        return pricing

    pricing.sheet_state = "hidden"
    multi_year.sheet_state = "visible"
    year_columns = ("P", "Q", "R", "S", "T")
    for year_number, column in enumerate(year_columns, start=1):
        multi_year.column_dimensions[column].hidden = year_number > request.contract_years

    last_year_column = year_columns[request.contract_years - 1]
    for row in (69, 71, 73, 75, 77, 78, 79):
        multi_year[f"V{row}"] = f"=SUM(P{row}:{last_year_column}{row})"

    multi_year.print_area = "A1:V80"
    multi_year.sheet_properties.pageSetUpPr.fitToPage = True
    multi_year.page_setup.orientation = "landscape"
    multi_year.page_setup.fitToWidth = 1
    multi_year.page_setup.fitToHeight = 1
    _select_only(workbook, multi_year)
    return multi_year


def _prepare_multi_year_locations(workbook, request: QuoteRequest) -> None:
    source = _prepare_multi_year(workbook, request)
    source.title = "_pricing_template"
    location_sheets = [source]
    for _ in request.locations[1:]:
        location_sheets.append(_copy_worksheet(workbook, source))

    for worksheet in list(workbook.worksheets):
        if worksheet not in location_sheets:
            workbook.remove(worksheet)

    titles = _location_sheet_titles(workbook, location_sheets, request.locations)
    customer_title = _customer_title(request)
    for worksheet, spec, title in zip(
        location_sheets,
        request.locations,
        titles,
        strict=True,
    ):
        worksheet.title = title
        display_name = _location_display_name(spec)
        _set_title_cell(
            worksheet,
            f"{customer_title} | Location - {display_name}",
        )
        worksheet.sheet_state = "visible"

    _select_only(workbook, location_sheets[0])


def _prepare_multi_location(workbook, request: QuoteRequest) -> None:
    summary = workbook["Cost Summary"]
    _set_title_cell(summary, _customer_title(request))

    required_specs = [location for location in request.locations if not location.optional]
    optional_specs = [location for location in request.locations if location.optional]

    required_sheets = [workbook[f"Location {index}"] for index in range(1, 7)]
    required_totals = ["P53", "P55", "P73", "P73", "P73", "P73"]
    required_subcontractors = ["P37", "P33", "P36", "P36", "P36", "P36"]
    required_materials = ["P42", "P44", "P62", "P62", "P62", "P62"]

    required_template = required_sheets[-1]
    for _ in range(len(required_sheets), len(required_specs)):
        required_sheets.append(_copy_worksheet(workbook, required_template))
        required_totals.append("P73")
        required_subcontractors.append("P36")
        required_materials.append("P62")

    active_required: list[tuple[object, str, str, str]] = []
    for index, worksheet in enumerate(required_sheets):
        if index < len(required_specs):
            worksheet.sheet_state = "visible"
            active_required.append(
                (
                    worksheet,
                    required_totals[index],
                    required_subcontractors[index],
                    required_materials[index],
                )
            )
        else:
            workbook.remove(worksheet)

    option_template = workbook["Option 1"]
    option_sheets = [option_template]
    for option_index in range(2, len(optional_specs) + 1):
        option_sheets.append(_copy_worksheet(workbook, option_template))

    active_options: list[object] = []
    for index, worksheet in enumerate(option_sheets):
        if index < len(optional_specs):
            worksheet.sheet_state = "visible"
            active_options.append(worksheet)
        else:
            workbook.remove(worksheet)

    active_sheets = [entry[0] for entry in active_required] + active_options
    active_specs = required_specs + optional_specs
    titles = _location_sheet_titles(workbook, active_sheets, active_specs)
    for worksheet, spec, title in zip(active_sheets, active_specs, titles, strict=True):
        worksheet.title = title
        _set_location_name(worksheet, _location_display_name(spec))
    workbook._sheets = [summary, *active_sheets]  # noqa: SLF001 - tab ordering

    (
        required_start,
        required_capacity,
        required_subtotal,
        optional_start,
        optional_capacity,
        optional_subtotal,
    ) = _expand_summary_rows(summary, len(required_specs), len(optional_specs))

    _populate_required_summary(
        summary,
        active_required,
        required_start,
        required_capacity,
        required_subtotal,
    )
    _populate_optional_summary(
        summary,
        active_options,
        optional_start,
        optional_capacity,
        optional_subtotal,
    )
    _repair_summary_totals(summary, required_subtotal, optional_subtotal)

    subcontractor_references = [
        f"{_quoted_sheet_name(worksheet.title)}!{subtotal}"
        for worksheet, _, subtotal, _ in active_required
    ] + [f"{_quoted_sheet_name(worksheet.title)}!P36" for worksheet in active_options]
    material_references = [
        f"{_quoted_sheet_name(worksheet.title)}!{subtotal}"
        for worksheet, _, _, subtotal in active_required
    ] + [f"{_quoted_sheet_name(worksheet.title)}!P62" for worksheet in active_options]
    summary["G9"] = _sum_formula(subcontractor_references)
    summary["G10"] = _sum_formula(material_references)

    summary.print_area = f"A1:J{optional_subtotal}"
    summary.sheet_properties.pageSetUpPr.fitToPage = True
    summary.page_setup.fitToWidth = 1
    summary.page_setup.fitToHeight = 1
    summary.sheet_state = "visible"
    _select_only(workbook, summary)


def _expand_summary_rows(
    worksheet,
    required_count: int,
    optional_count: int,
) -> tuple[int, int, int, int, int, int]:
    required_capacity = max(BASE_REQUIRED_ROW_COUNT, required_count)
    required_extra = required_capacity - BASE_REQUIRED_ROW_COUNT
    if required_extra:
        _insert_styled_rows(
            worksheet,
            BASE_REQUIRED_SUBTOTAL_ROW,
            required_extra,
            BASE_REQUIRED_SUBTOTAL_ROW - 1,
        )

    optional_capacity = max(BASE_OPTION_ROW_COUNT, optional_count)
    optional_extra = optional_capacity - BASE_OPTION_ROW_COUNT
    option_subtotal_before_insert = BASE_OPTION_SUBTOTAL_ROW + required_extra
    if optional_extra:
        _insert_styled_rows(
            worksheet,
            option_subtotal_before_insert,
            optional_extra,
            option_subtotal_before_insert - 1,
        )

    required_subtotal = BASE_REQUIRED_SUBTOTAL_ROW + required_extra
    optional_start = BASE_OPTION_START_ROW + required_extra
    optional_subtotal = option_subtotal_before_insert + optional_extra
    return (
        REQUIRED_START_ROW,
        required_capacity,
        required_subtotal,
        optional_start,
        optional_capacity,
        optional_subtotal,
    )


def _insert_styled_rows(
    worksheet,
    insertion_row: int,
    amount: int,
    source_row: int,
) -> None:
    formula_cells = [
        (cell.column, cell.row, cell.value)
        for row in worksheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    ]
    shifted_row_dimensions = {
        row_index: copy(dimension)
        for row_index, dimension in worksheet.row_dimensions.items()
        if row_index >= insertion_row
    }
    for row_index in shifted_row_dimensions:
        del worksheet.row_dimensions[row_index]

    affected_ranges: list[tuple[int, int, int, int]] = []
    for merged_range in list(worksheet.merged_cells.ranges):
        if merged_range.max_row >= insertion_row:
            affected_ranges.append(
                (
                    merged_range.min_row,
                    merged_range.max_row,
                    merged_range.min_col,
                    merged_range.max_col,
                )
            )
            worksheet.unmerge_cells(str(merged_range))

    worksheet.insert_rows(insertion_row, amount=amount)

    for column, original_row, formula in formula_cells:
        target_row = (
            original_row + amount
            if original_row >= insertion_row
            else original_row
        )
        worksheet.cell(target_row, column).value = _shift_formula_rows(
            formula,
            insertion_row,
            amount,
        )

    for original_row, dimension in shifted_row_dimensions.items():
        target_row = original_row + amount
        dimension.index = target_row
        worksheet.row_dimensions[target_row] = dimension

    _shift_data_validation_rows(
        worksheet,
        insertion_row,
        amount,
    )

    for min_row, max_row, min_col, max_col in affected_ranges:
        if min_row >= insertion_row:
            min_row += amount
        max_row += amount
        worksheet.merge_cells(
            start_row=min_row,
            end_row=max_row,
            start_column=min_col,
            end_column=max_col,
        )

    for target_row in range(insertion_row, insertion_row + amount):
        worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
        for column in range(1, 11):
            source = worksheet.cell(source_row, column)
            target = worksheet.cell(target_row, column)
            target._style = copy(source._style)
            if source.has_style:
                target.number_format = source.number_format
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)


def _populate_required_summary(
    worksheet,
    active_required,
    start_row: int,
    capacity: int,
    subtotal_row: int,
) -> None:
    for index, row in enumerate(range(start_row, start_row + capacity)):
        worksheet.row_dimensions[row].hidden = index >= len(active_required)
        for column in (1, 2, 3, 4, 5, 7):
            worksheet.cell(row, column).value = None
        worksheet[f"F{row}"] = f"=C{row}*D{row}*E{row}"
        worksheet[f"H{row}"] = f"=(F{row}*G{row})"
        worksheet[f"I{row}"] = 1
        worksheet[f"J{row}"] = f"=H{row}/I{row}"

        if index < len(active_required):
            location_sheet, total_cell, _, _ = active_required[index]
            worksheet[f"A{row}"] = index + 1
            worksheet[f"B{row}"] = f"={_quoted_sheet_name(location_sheet.title)}!R1"
            worksheet[f"C{row}"] = 1
            worksheet[f"D{row}"] = 1
            worksheet[f"E{row}"] = 1
            worksheet[f"G{row}"] = (
                f"={_quoted_sheet_name(location_sheet.title)}!{total_cell}"
            )

    worksheet[f"H{subtotal_row}"] = f"=SUM(H{start_row}:H{subtotal_row - 1})"
    worksheet[f"J{subtotal_row}"] = f"=SUM(J{start_row}:J{subtotal_row - 1})"


def _populate_optional_summary(
    worksheet,
    option_sheets,
    start_row: int,
    capacity: int,
    subtotal_row: int,
) -> None:
    for index, row in enumerate(range(start_row, start_row + capacity)):
        worksheet.row_dimensions[row].hidden = index >= len(option_sheets)
        for column in (1, 2, 3, 4, 5, 7):
            worksheet.cell(row, column).value = None
        worksheet[f"F{row}"] = f"=C{row}*D{row}*E{row}"
        worksheet[f"H{row}"] = f"=F{row}*G{row}"
        worksheet[f"I{row}"] = 1
        worksheet[f"J{row}"] = f"=H{row}/I{row}"

        if index < len(option_sheets):
            option_sheet = option_sheets[index]
            worksheet[f"A{row}"] = index + 1
            worksheet[f"B{row}"] = f"={_quoted_sheet_name(option_sheet.title)}!R1"
            worksheet[f"C{row}"] = 1
            worksheet[f"D{row}"] = 1
            worksheet[f"E{row}"] = 1
            worksheet[f"G{row}"] = f"={_quoted_sheet_name(option_sheet.title)}!P73"

    worksheet[f"H{subtotal_row}"] = f"=SUM(H{start_row}:H{subtotal_row - 1})"
    worksheet[f"J{subtotal_row}"] = f"=SUM(J{start_row}:J{subtotal_row - 1})"


def _repair_summary_totals(
    worksheet,
    required_subtotal: int,
    optional_subtotal: int,
) -> None:
    # Project Subtotal (J35), Project Options Subtotal (J43), Project Total
    # (J45), HST (J46), TOTAL (J47), the HST check (J48), and the Ainsworth
    # price (J53) already reference the required/optional subtotal rows in
    # the template and are kept correct automatically: row insertion in
    # _insert_styled_rows() rewrites every formula's row references
    # workbook-wide, including these, whether or not a row shift happened.
    worksheet[f"A{required_subtotal}"] = "Subtotal Required Locations"
    worksheet[f"B{required_subtotal + 2}"] = "Optional Location"
    worksheet[f"A{optional_subtotal}"] = "Subtotal Optional Locations"
    worksheet["G7"] = (
        "=J6+J8+J9+J10+J11+J12+J16+"
        f"J{required_subtotal}+J{optional_subtotal}"
    )


def _copy_worksheet(workbook, source):
    clone = workbook.copy_worksheet(source)
    clone.data_validations = deepcopy(source.data_validations)
    if source.print_area:
        local_print_area = str(source.print_area).split("!", 1)[-1].replace("$", "")
        clone.print_area = local_print_area
    return clone


def _location_sheet_titles(workbook, worksheets, specs) -> list[str]:
    active_ids = {id(worksheet) for worksheet in worksheets}
    reserved = {
        title.casefold()
        for worksheet, title in zip(
            workbook.worksheets,
            workbook.sheetnames,
            strict=True,
        )
        if id(worksheet) not in active_ids
    }
    reserved.add(LABOUR_RATES_SHEET.casefold())

    titles: list[str] = []
    for spec in specs:
        suffix = " (Optional)" if spec.optional else ""
        base = INVALID_SHEET_TITLE.sub("-", spec.name).strip().strip("'")
        base = re.sub(r"\s+", " ", base) or "Location"
        counter = 1
        while True:
            duplicate_suffix = "" if counter == 1 else f" ({counter})"
            available = 31 - len(duplicate_suffix) - len(suffix)
            truncated_base = _truncate_utf16(base, available).rstrip()
            candidate = f"{truncated_base}{duplicate_suffix}{suffix}"
            if candidate.casefold() not in reserved:
                break
            counter += 1
        reserved.add(candidate.casefold())
        titles.append(candidate)

    occupied = {title.casefold() for title in workbook.sheetnames}
    desired = {title.casefold() for title in titles}
    for index, worksheet in enumerate(worksheets, start=1):
        counter = index
        while True:
            temporary = f"_erics_temp_{counter:02d}"
            if temporary.casefold() not in occupied | desired:
                break
            counter += len(worksheets)
        occupied.discard(worksheet.title.casefold())
        worksheet.title = temporary
        occupied.add(temporary.casefold())
    return titles


def _location_display_name(spec: LocationSpec) -> str:
    return f"{spec.name} (Optional)" if spec.optional else spec.name


def _truncate_utf16(value: str, maximum_units: int) -> str:
    """Truncate text to Excel's UTF-16-based worksheet-title limit."""
    units = 0
    result: list[str] = []
    for character in value:
        character_units = len(character.encode("utf-16-le")) // 2
        if units + character_units > maximum_units:
            break
        result.append(character)
        units += character_units
    return "".join(result)


def _quoted_sheet_name(title: str) -> str:
    return f"'{title.replace(chr(39), chr(39) * 2)}'"


def _add_labour_rates_reference(
    workbook,
    settings: LabourRateSettings,
) -> None:
    worksheet = workbook.create_sheet(LABOUR_RATES_SHEET)
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "B6"
    worksheet.sheet_properties.tabColor = "086A9D"

    navy = "173F60"
    blue = "086A9D"
    pale_blue = "E8F4FA"
    pale_gray = "F0F4F7"
    border_color = "D6E0E6"
    white = "FFFFFF"
    text_color = "142A3A"
    muted = "60717E"
    thin = Side(style="thin", color=border_color)

    worksheet.merge_cells("A1:N1")
    worksheet["A1"] = f"GDI Ainsworth Labour Rates — {settings.year}"
    worksheet["A1"].font = Font(name="Segoe UI", size=20, bold=True, color=white)
    worksheet["A1"].fill = PatternFill("solid", fgColor=navy)
    worksheet["A1"].alignment = Alignment(vertical="center")
    worksheet.row_dimensions[1].height = 34

    worksheet.merge_cells("A2:N2")
    worksheet["A2"] = (
        "Annual reference used by every pricing sheet in this workbook. "
        "Manage these values in Power Systems Costing Workbook under Menu > Labour rates."
    )
    worksheet["A2"].font = Font(name="Segoe UI", size=10, color=muted)
    worksheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    worksheet.row_dimensions[2].height = 30

    groups = (
        ("B4:D4", "Cost rates"),
        ("E4:G4", "Cost rates with truck"),
        ("I4:K4", "Sell rates"),
        ("L4:N4", "Sell rates with truck"),
    )
    for cell_range, label in groups:
        worksheet.merge_cells(cell_range)
        cell = worksheet[cell_range.split(":", 1)[0]]
        cell.value = label
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.font = Font(name="Segoe UI", size=10, bold=True, color=white)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    headers = (
        "Classification",
        "Regular",
        "Double time",
        "Shift 20%",
        "Regular",
        "Double time",
        "Shift 20%",
        "Margin",
        "Regular",
        "Double time",
        "Shift 20%",
        "Regular",
        "Double time",
        "Shift 20%",
    )
    for column, label in enumerate(headers, start=1):
        cell = worksheet.cell(5, column, label)
        cell.fill = PatternFill("solid", fgColor=pale_blue)
        cell.font = Font(name="Segoe UI", size=9, bold=True, color=text_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    worksheet.row_dimensions[5].height = 30

    for index, rate in enumerate(settings.rates, start=6):
        worksheet.cell(index, 1, rate.classification)
        worksheet.cell(index, 2, rate.regular)
        worksheet.cell(index, 3, rate.double_time)
        worksheet.cell(index, 4, rate.shift_20)
        worksheet.cell(index, 5, f"=B{index}+$B$19")
        worksheet.cell(index, 6, f"=C{index}+$B$19")
        worksheet.cell(index, 7, f"=D{index}+$B$19")
        worksheet.cell(index, 8, "=$B$20")
        worksheet.cell(index, 9, f"=B{index}/H{index}")
        worksheet.cell(index, 10, f"=C{index}/H{index}")
        worksheet.cell(index, 11, f"=D{index}/H{index}")
        worksheet.cell(index, 12, f"=E{index}/H{index}")
        worksheet.cell(index, 13, f"=F{index}/H{index}")
        worksheet.cell(index, 14, f"=G{index}/H{index}")
        fill = PatternFill("solid", fgColor=white if index % 2 == 0 else pale_gray)
        for column in range(1, 15):
            cell = worksheet.cell(index, column)
            cell.fill = fill
            cell.border = Border(bottom=thin)
            cell.font = Font(name="Segoe UI", size=9, color=text_color)
            cell.alignment = Alignment(
                horizontal="left" if column == 1 else "right",
                vertical="center",
            )
            if 2 <= column <= 7 or 9 <= column <= 14:
                cell.number_format = '$0.00'
            elif column == 8:
                cell.number_format = '0%'

    worksheet["A17"] = "Annual settings"
    worksheet["A17"].font = Font(name="Segoe UI", size=12, bold=True, color=navy)
    annual_settings = (
        (18, "Rate year", settings.year, "0"),
        (19, "Truck premium", settings.truck_premium, '$0.00'),
        (20, "Sell margin", settings.sell_margin, '0%'),
    )
    for row, label, value, number_format in annual_settings:
        worksheet.cell(row, 1, label)
        worksheet.cell(row, 2, value)
        for column in (1, 2):
            worksheet.cell(row, column).fill = PatternFill("solid", fgColor=pale_gray)
            worksheet.cell(row, column).border = Border(bottom=thin)
            worksheet.cell(row, column).font = Font(
                name="Segoe UI",
                size=9,
                bold=column == 1,
                color=text_color,
            )
        worksheet.cell(row, 2).number_format = number_format

    worksheet.auto_filter.ref = f"A5:N{5 + len(settings.rates)}"
    worksheet.column_dimensions["A"].width = 24
    for column in range(2, 15):
        worksheet.column_dimensions[get_column_letter(column)].width = 13
    worksheet.print_area = "A1:N20"
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1


def _link_labour_rate_tables(workbook, settings: LabourRateSettings) -> None:
    rate_rows = {
        rate.classification.casefold(): row
        for row, rate in enumerate(settings.rates, start=6)
    }
    reference = _quoted_sheet_name(LABOUR_RATES_SHEET)
    cost_columns = {3: "B", 4: "E", 6: "C", 7: "F", 8: "D", 9: "G"}
    sell_columns = {3: "I", 4: "L", 6: "J", 7: "M", 8: "K", 9: "N"}

    for worksheet in workbook.worksheets:
        if worksheet.title == LABOUR_RATES_SHEET:
            continue
        cost_header = None
        sell_header = None
        for row in range(1, worksheet.max_row + 1):
            label = _normalized_rate_label(worksheet.cell(row, 2).value)
            if label == "aps (cost)":
                cost_header = row
            elif label == "aps (sell)":
                sell_header = row

        if cost_header is not None:
            _link_rate_section(
                worksheet,
                cost_header,
                rate_rows,
                reference,
                cost_columns,
            )
        if sell_header is not None:
            _link_rate_section(
                worksheet,
                sell_header,
                rate_rows,
                reference,
                sell_columns,
            )
            worksheet.cell(sell_header + 1, 11).value = f"={reference}!$H$6"


def _link_rate_section(
    worksheet,
    header_row: int,
    rate_rows: dict[str, int],
    reference: str,
    column_mapping: dict[int, str],
) -> None:
    for row in range(header_row + 1, min(header_row + 12, worksheet.max_row) + 1):
        label = _normalized_rate_label(worksheet.cell(row, 2).value)
        rate_row = rate_rows.get(label)
        if rate_row is None:
            if label:
                break
            continue
        for target_column, source_column in column_mapping.items():
            worksheet.cell(row, target_column).value = (
                f"={reference}!${source_column}${rate_row}"
            )


def _normalized_rate_label(value: object) -> str:
    if not isinstance(value, str):
        return ""
    return " ".join(value.split()).casefold()


def _shift_formula_rows(formula: str, insertion_row: int, amount: int) -> str:
    """Apply Excel-style structural row insertion to local A1 references."""

    def replace_reference(match: re.Match[str]) -> str:
        row = int(match.group("row"))
        if row < insertion_row:
            return match.group(0)
        return (
            f"{match.group('column')}{match.group('absolute_row')}"
            f"{row + amount}"
        )

    return CELL_REFERENCE_PATTERN.sub(replace_reference, formula)


def _shift_data_validation_rows(worksheet, insertion_row: int, amount: int) -> None:
    for validation in worksheet.data_validations.dataValidation:
        shifted_ranges: list[str] = []
        for cell_range in validation.ranges.ranges:
            min_column = get_column_letter(cell_range.min_col)
            max_column = get_column_letter(cell_range.max_col)
            min_row = cell_range.min_row
            max_row = cell_range.max_row
            if min_row >= insertion_row:
                min_row += amount
                max_row += amount
            elif max_row >= insertion_row:
                max_row += amount
            shifted_ranges.append(
                f"{min_column}{min_row}:"
                f"{max_column}{max_row}"
                if (
                    cell_range.min_col != cell_range.max_col
                    or min_row != max_row
                )
                else f"{min_column}{min_row}"
            )
        validation.sqref = " ".join(shifted_ranges)
        if isinstance(validation.formula1, str) and validation.formula1.startswith("="):
            validation.formula1 = _shift_formula_rows(
                validation.formula1,
                insertion_row,
                amount,
            )


def _set_location_name(worksheet, name: str) -> None:
    _set_title_cell(worksheet, name)
    _set_literal_text(worksheet["R1"], name)


def _set_title_cell(worksheet, title: str) -> None:
    cell = worksheet["A1"]
    if isinstance(cell, MergedCell):
        raise WorkbookGenerationError(f"Cannot update the title on {worksheet.title}.")
    _set_literal_text(cell, title)
    alignment = copy(cell.alignment)
    alignment.wrap_text = True
    alignment.shrink_to_fit = True
    cell.alignment = alignment


def _set_literal_text(cell, value: str) -> None:
    """Store user-controlled text without letting Excel interpret it as a formula."""
    cell.value = value
    cell.data_type = "s"


def _customer_title(request: QuoteRequest) -> str:
    details = request.customer_name
    if request.project_scope:
        details = f"{details} | {request.project_scope}"
    return f"Customer and Project Scope - {details}"


def _sum_formula(references: list[str]) -> str:
    if not references:
        return "=0"
    return f"=SUM({','.join(references)})"


def _request_full_recalculation(workbook) -> None:
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True


def _sanitize_workbook_metadata(workbook) -> None:
    """Remove personal source-template metadata from every generated workbook."""
    workbook.properties.creator = "GDI Ainsworth"
    workbook.properties.lastModifiedBy = "GDI Ainsworth"
    workbook.properties.lastPrinted = None


def _select_only(workbook, selected_worksheet) -> None:
    """Prevent source-template tab grouping from leaking into generated files."""
    for worksheet in workbook.worksheets:
        worksheet.sheet_view.tabSelected = False
    selected_worksheet.sheet_view.tabSelected = True
    selected_index = workbook.sheetnames.index(selected_worksheet.title)
    workbook.active = selected_index
    if workbook.views:
        workbook.views[0].activeTab = selected_index


def _unique_destination(request: QuoteRequest, output_directory: Path) -> Path:
    if request.mode is QuoteMode.MULTI_LOCATION:
        mode_name = "Multi Location"
    elif request.contract_years == 1:
        mode_name = "Multi Location Pricing" if request.include_locations else "Pricing"
    else:
        mode_name = f"{request.contract_years} Year Contract"
        if request.include_locations:
            mode_name = f"{mode_name} - Multi Location"
    project_part = f" - {request.project_scope}" if request.project_scope else ""
    base = _safe_filename(f"{request.customer_name}{project_part} - {mode_name}")
    candidate = output_directory / f"{base}.xlsx"
    suffix = 2
    while True:
        try:
            descriptor = os.open(
                candidate,
                os.O_CREAT | os.O_EXCL | os.O_WRONLY,
            )
        except FileExistsError:
            candidate = output_directory / f"{base} ({suffix}).xlsx"
            suffix += 1
            continue
        os.close(descriptor)
        return candidate


def _safe_filename(value: str) -> str:
    value = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "-", value)
    value = re.sub(r"\s+", " ", value).strip(" .")
    return value[:150] or "Quote"
